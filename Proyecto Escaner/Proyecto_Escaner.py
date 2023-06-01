import tkinter as tk
from tkinter import messagebox

import face_recognition as fr
from PIL import ImageTk, Image
from pathlib import Path
import io
import os
from cv2 import cv2
from datetime import datetime
import numpy
import pygame
import locale

#LIBRERIAS PARA EL EXCEL
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo




#----------PARA ESTABLECER EL IDIOMA EN ESPAÑOL
locale.setlocale(locale.LC_TIME, 'es_ES')


#-----------FUNCION QUE CODIFICA LAS IMAGENES DE LOS EMPLEADOS PARA PODER COMPARARLAS-----------
def codificar(imagenes):
    lista_codificados = []
    for imagen in imagenes:
        imagen_codificada = fr.face_encodings(imagen)[0]
        lista_codificados.append(imagen_codificada)
    return lista_codificados


#--------SE CONSTRUYE UNA FUNCION ENCARGADA DE REGISTRAR LA ASISTENCIA
def asistencia(nombre):
    f= open('Registro_de_asistencia/asistencia.csv', 'r+')
    contenido= f.readlines()
    nombres_marcados= []
    nombre= " "+nombre

    for linea in contenido:
        try:
            nombres_marcados.append(linea.split(',')[2])
        except:
            pass

    if nombre not in nombres_marcados:
        puntualidad = "Tarde" if int(datetime.today().strftime("%H"))>8 else "Temprano"
        dia_actual = datetime.today().strftime("%A")
        fecha_actual = datetime.today().strftime("%d/%B/%Y")
        hora_actual = datetime.today().strftime("%Hh:%Mm:%Hs")
        f.write(f'\n{fecha_actual}, {dia_actual},{nombre}, {hora_actual}, {puntualidad}')


    #----------PARA CREAR Y GUARDAR EN UN ARCHIVO EXCEL----------
    # Leer el archivo CSV
    df = pd.read_csv('Registro_de_asistencia/asistencia.csv')
    df = df.astype(str)


    # Crear un nuevo archivo Excel
    wb = Workbook()
    ws = wb.active

    # Escribir los datos en el archivo Excel
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # Definir el estilo de la cabecera
    fill = PatternFill(start_color="50A2EB", end_color="FFFFFF", fill_type="solid")
    for cell in ws[1]:
        cell.fill = fill

    # Agregar una tabla dinámica
    tablas = Table(displayName="TablaDatos", ref=f'A1:E{df.shape[0] + 1}')
    tablas.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
                                           showRowStripes=True, showColumnStripes=False)
    ws.add_table(tablas)

    # Guardar el archivo Excel
    try:
        wb.save('Registro_de_asistencia/asistencia.xlsx')
    except:
        messagebox.showerror("Error", "Usted tiene abierto el archivo excel\nno se guardaran los cambios")

    f.close()


#---------Construye nuestro propio MESSAGEBOX---------
class PopupWindow(tk.Toplevel):
    def __init__(self, parent, message, timeout):
        super().__init__(parent)
        self.title("Información")
        self.geometry("300x100")
        self.resizable(False, False)
        self.protocol("WM_DELETE_WINDOW", self.close_window)

        # Establecer el icono de la ventana
        self.iconbitmap("Imagenes_del_programa/favicon.ico")

        self.label_message = tk.Label(self, text=message)
        self.label_message.pack(pady=20)

        self.after(timeout, self.close_window)

    def close_window(self):
        self.destroy()



class Interface:

    #PARA CREAR LA INTERFAZ DEL MENU
    def __init__(self, root):
        self.root = root
        self.root.title("HUALKANA SAC - SENATI")

        # Configurar el tamaño de la ventana
        self.root.geometry("480x360")
        self.root.resizable(0, 0)

        # Cambiar el favicon
        favicon_path = "Imagenes_del_programa/favicon.ico"  # Ruta del archivo de icono (formato .ico)
        self.root.iconbitmap(favicon_path)


        # Establecer el color de fondo
        bg_color = (255, 255, 255)  # Blanco en RGB
        bg_hex = '#%02x%02x%02x' % bg_color
        bg_hex_titulo= '#%02x%02x%02x' % (15,105,131)
        bg_hex_fondo = '#%02x%02x%02x' % (137,221,245)
        bg_hex_boton = '#%02x%02x%02x' % (167,178,185)
        self.root.config(bg=bg_hex_fondo)

        # Etiqueta para el título
        self.label_title = tk.Label(root, text="HUALKANA SAC - SENATI", font=("Arial", 24),bg=bg_hex)
        self.label_title.pack(pady=30)

        # Marco para los botones
        self.button_frame = tk.Frame(root,bg=bg_hex_fondo)
        self.button_frame.pack()

        # Botón para iniciar la cámara web
        self.button_start = tk.Button(self.button_frame, text="Iniciar cámara", command=self.start_webcam,bg=bg_hex_boton)
        self.button_start.pack(side=tk.LEFT, padx=6)

        # Botón para salir
        self.button_quit = tk.Button(self.button_frame, text="Salir", command=self.root.quit,bg=bg_hex_boton)
        self.button_quit.pack(side=tk.LEFT, padx=6)

        # Marco para las imágenes
        self.image_frame = tk.Frame(root,bg=bg_hex_fondo)
        self.image_frame.pack(pady=20)  # Añadir espacio vertical entre los botones y las imágenes

        # Cargar las imágenes
        image1 = Image.open("Imagenes_del_programa/logo_senati.png")  # Ruta de la primera imagen
        image1 = image1.resize((84, 84))  # Cambiar el tamaño si es necesario
        self.photo1 = ImageTk.PhotoImage(image1)
        self.label_image1 = tk.Label(self.image_frame, image=self.photo1,bg=bg_hex)
        self.label_image1.pack(side=tk.LEFT, padx=10)  # Añadir espacio horizontal entre las imágenes

        image2 = Image.open("Imagenes_del_programa/logo_hualkana.png")  # Ruta de la segunda imagen
        image2 = image2.resize((107, 84))  # Cambiar el tamaño si es necesario
        self.photo2 = ImageTk.PhotoImage(image2)
        self.label_image2 = tk.Label(self.image_frame, image=self.photo2,bg=bg_hex)
        self.label_image2.pack(side=tk.LEFT, padx=10)  # Añadir espacio horizontal entre las imágenes

        # Etiqueta para el comentario
        self.label_comment = tk.Label(root,
                                      text="* Escáner biométrico facial desarrollado por los alumnos de Senati:\n\t- Victorio Zanabria Kevin Dax\n\t- Zanabria Hurtado Yoselyn Patricia",
                                      font=("Arial", 10), bg=bg_hex_fondo,justify="left")
        self.label_comment.pack(pady=15)

        self.cap = None

    #INICIAR LA CAMARA WEB Y SUS PROCESOS
    def start_webcam(self):

        #MESSAGEBOX QUE DICE QUE ESPERE
        if True:
            self.root.withdraw()
            cv2.waitKey(500)
            self.show_popup("Espere", "La camara esta inicializando...\nESPERE, POR FAVOR...",1000)


        # --------AQUI SE CONSTRUYE LA RUTA Y LAS IMAGENES QUE USARA EL PROGRAMA PARA IDENTIFICAR AL PERSONAL------------
        if True:
            ruta = 'Empleados'
            empleados_nombres = []
            empleados_imagenes = []

            for carpeta, _, archivos in os.walk(ruta):
                for archivo in archivos:
                    archivo_actual = cv2.imread(f'{carpeta}/{archivo}')
                    empleados_imagenes.append(archivo_actual)
                    empleados_nombres.append(os.path.splitext(archivo)[0])

            empleados_codificados = codificar(empleados_imagenes)


            # ---------AQUI SE TOMA LA INSTANTANEA DEL VIDEO
            captura = cv2.VideoCapture(0, cv2.CAP_DSHOW)  # Usando la camara DirectShow
            captura.set(cv2.CAP_PROP_FPS, 60)
            captura.set(cv2.CAP_PROP_FRAME_WIDTH, 480)
            captura.set(cv2.CAP_PROP_FRAME_HEIGHT, 360)

            # MESSAGEBOX QUE DICE QUE ESPERE
            self.show_popup("Espere", "La camara esta inicializando...\nESPERE, POR FAVOR...",2000)

        # ---------AQUI SE CONSTRUYE EL PROCESO EN DONDE SE IDENTIFICA A LA PERSONA---------------------
        running = True
        ventana_estado = 1.0
        while running:
            exito, instantanea = captura.read()

            #SI EL PROGRAMA NO DETECTA UNA CAMARA EN EL SISTEMA
            if not exito:
                messagebox.showerror("Error", "No se pudo abrir la cámara web.")
                running = False

            #SI EL PROGRAMA SI DETECTA UNA CAMARA EN EL SISTEMA
            else:
                #TOMA UNA INSTANTANEA Y LA PROCESA
                if True:
                    altura, ancho, profundidad = instantanea.shape
                    cara_instantanea = fr.face_locations(instantanea)
                    codificar_instantanea = fr.face_encodings(instantanea, cara_instantanea)

                #SI NO HAY UNA CARA EN LA CAMARA
                if cara_instantanea == []:
                    pygame.mixer_music.stop()


                #SI HAY UNA CARA EN LA CAMARA
                else:
                    for caraubic, caracodif in zip(cara_instantanea, codificar_instantanea):
                        distancias = fr.face_distance(empleados_codificados, caracodif)
                        verificaciones = fr.compare_faces(empleados_codificados, caracodif)

                        indice_coincidencia = numpy.argmin(distancias)

                        #REPRODUCE EL SONIDO DE LA ALARMA SI LA CARA NO PERTENECE A LA EMPRESA
                        if distancias[indice_coincidencia] > 0.6:
                            #SI LA MUSICA NO ESTA OCUPADA
                            if not pygame.mixer.music.get_busy():
                                pygame.mixer.music.load('Sonidos/noEmpleado.mp3')
                                pygame.mixer.music.set_volume(0.8)
                                pygame.event.poll()
                                pygame.mixer.music.play(-1)

                        # REPRODUCE UN SONIDO SI LA CARA SIPERTENECE A LA EMPRESA
                        else:

                            #SONIDO QUE AVISA QUE EL EMPLEADO PERTENECE A LA EMPRESA
                            if not pygame.mixer.music.get_busy():
                                pygame.mixer.music.load('Sonidos/siEmpleado.mp3')
                                pygame.mixer.music.set_volume(0.8)
                                pygame.event.poll()
                                pygame.mixer.music.play(1)

                            #NOMBRE Y UBISCACION DE LAS COORDENADAS DE LA CARA
                            nombre = empleados_nombres[indice_coincidencia]
                            y1, x2, y2, x1 = caraubic

                            #REGISTRA LA ASISTENCIA EN EL ARCHIVO EXCEL
                            asistencia(nombre)

                            #------------ESTO ES PARA DARLE RECONOCIMIENTO A LA IMAGEN Y LO MUESRE EN LA CAMARA---------
                            if True:
                                #cv2.putText(instantanea, 'BIENVENIDO A HUALKANA', (int(ancho*36 / 100) , 15), cv2.FONT_HERSHEY_COMPLEX, 0.5, (0, 255, 0), 1)       #MUESTRA UN MENSAJE EN LA CAMARA
                                cv2.rectangle(instantanea, (x1, y1), (x2, y2), (0, 255, 0), 1)                                                         #MUESTRA UN RECTANGULO EN EL ROSTRO
                                #cv2.rectangle(instantanea, (x1-15, y2 - 20), (x2+15, y2+1), (0, 255, 0), cv2.FILLED)                                   #MUESTRA UN BLOQUE QUE RODEA AL NOMBRE
                                cv2.putText(instantanea, nombre, (x1 - 50, y2 + 15), cv2.FONT_HERSHEY_COMPLEX, 0.5,(255, 0, 0), 1)                      #MUESTRA EL NOMBRE
                                pass

                #PARA CERRAR EL PROGRAMA
                cv2.imshow('CAMARA WEB', instantanea)
                if cv2.waitKey(1) & 0xFF == ord('q'):
                    running = False


                #PARA CERRAR EL PROGRAMA
                if ventana_estado <= 0:
                    running = False

                ventana_estado = cv2.getWindowProperty("CAMARA WEB", cv2.WND_PROP_VISIBLE)

        # ---------------------ESTO CIERRA LAS VENTANAS USADAS POR NUESTRO PROGRAMA--------------------
        captura.release()  # cierra la camara web y libera a la variable captura
        cv2.destroyAllWindows()  # Cierra todas las ventanas abiertas por OpenCV
        try:
            self.close_window()
        except:
            pass

    #CREAR NUESTRO PROPIO MESSAGEBOX
    def show_popup(self, title, message,time):
        popup = PopupWindow(self.root, message, time)
        popup.title(title)
        self.root.wait_window(popup)


    #CERRAR VENTANAS EMERGENTES Y CERRAR EL LOOP
    def close_window(self):
        self.root.destroy()

    #LOOP DEL PROGRAMA
    def run(self):
        self.root.mainloop()


pygame.init()
pygame.mixer.init()
pygame.mixer_music.load('Sonidos/siEmpleado.mp3')
pygame.mixer_music.stop()

root = tk.Tk()
interface = Interface(root)
interface.run()
pygame.quit()